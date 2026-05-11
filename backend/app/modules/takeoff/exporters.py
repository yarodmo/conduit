"""
Conduit Backend — Takeoff Exporters (M5)
Excel (openpyxl) + PDF (reportlab) with org branding.

Spec: SUPERA A TRIMBLE + PLANSWIFT
  Excel: subtotales por sistema + pestaña Supplier Contacts
  PDF: membrete org, resumen ejecutivo, tabla completa, firma del aprobador

Bliss Systems LLC — APEX Standard
"""

import io
from decimal import Decimal
from typing import Any


def export_excel(job: Any, items: list[Any], org_name: str) -> bytes:
    """
    Generate Excel takeoff report.
    Sheet 1: Takeoff Items (grouped by system with subtotals)
    Sheet 2: Supplier Contacts
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Takeoff Items ──────────────────────────────────────────
    ws = wb.active
    ws.title = "MEP Takeoff"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"MEP Takeoff Report — {org_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Job ID: {job.id}  |  Model: {job.model_version or 'N/A'}  |  Status: {job.status}"
    ws["A2"].font = Font(italic=True, size=10)

    # Headers
    headers = ["Type", "Tag", "Qty", "Unit", "Specification",
               "System", "CFM/GPM", "Unit Cost ($)", "Total Cost ($)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Group by system
    systems: dict[str, list[Any]] = {}
    for item in items:
        s = item.system or "Other"
        systems.setdefault(s, []).append(item)

    row = 5
    grand_total = Decimal("0")

    for system, sys_items in sorted(systems.items()):
        # System header
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = f"▶ {system.upper()}"
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].fill = subtotal_fill
        row += 1

        sys_total = Decimal("0")
        for item in sys_items:
            ws.cell(row=row, column=1, value=item.type)
            ws.cell(row=row, column=2, value=item.tag or "")
            ws.cell(row=row, column=3, value=float(item.quantity))
            ws.cell(row=row, column=4, value=item.unit)
            ws.cell(row=row, column=5, value=item.specification)
            ws.cell(row=row, column=6, value=item.system or "")
            ws.cell(row=row, column=7, value=float(item.cfm_or_gpm) if item.cfm_or_gpm else "")
            ws.cell(row=row, column=8, value=float(item.unit_cost_usd) if item.unit_cost_usd else "")
            ws.cell(row=row, column=9, value=float(item.total_cost_usd) if item.total_cost_usd else "")

            if item.total_cost_usd:
                sys_total += item.total_cost_usd

            # Flag low-confidence items
            if item.requires_review:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
            row += 1

        # System subtotal
        ws.cell(row=row, column=8, value="Subtotal").font = bold_font
        ws.cell(row=row, column=9, value=float(sys_total)).font = bold_font
        ws.cell(row=row, column=9).fill = subtotal_fill
        grand_total += sys_total
        row += 1

    # Grand total
    row += 1
    ws.cell(row=row, column=8, value="GRAND TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=9, value=float(grand_total)).font = Font(bold=True, size=12)
    ws.cell(row=row, column=9).fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    ws.cell(row=row, column=9).font = Font(bold=True, size=12, color="FFFFFF")

    # Column widths
    col_widths = [14, 12, 8, 8, 45, 14, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Supplier Contacts ──────────────────────────────────────
    ws2 = wb.create_sheet("Supplier Contacts")
    s_headers = ["Type", "Specification", "Supplier", "Contact", "Lead Days", "Unit Cost ($)"]
    for col, h in enumerate(s_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    s_row = 2
    seen_suppliers: set[str] = set()
    for item in items:
        if item.catalog_item and item.catalog_item.supplier_name:
            key = f"{item.type}:{item.catalog_item.supplier_name}"
            if key not in seen_suppliers:
                seen_suppliers.add(key)
                ws2.cell(row=s_row, column=1, value=item.type)
                ws2.cell(row=s_row, column=2, value=item.specification)
                ws2.cell(row=s_row, column=3, value=item.catalog_item.supplier_name)
                ws2.cell(row=s_row, column=4, value=item.catalog_item.supplier_contact or "")
                ws2.cell(row=s_row, column=5, value=item.catalog_item.supplier_lead_days or "")
                ws2.cell(row=s_row, column=6,
                         value=float(item.unit_cost_usd) if item.unit_cost_usd else "")
                s_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(
    job: Any,
    items: list[Any],
    org_name: str,
    approver_name: str | None = None,
) -> bytes:
    """
    Generate PDF takeoff report with org branding.
    Includes: summary, full table, low-confidence notes, approver signature line.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                             leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                             topMargin=0.75 * inch, bottomMargin=0.75 * inch)

    styles = getSampleStyleSheet()
    CONDUIT_BLUE = colors.HexColor("#1F4E79")

    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                  textColor=CONDUIT_BLUE, fontSize=16, spaceAfter=6)
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"],
                                     fontSize=9, textColor=colors.grey, spaceAfter=12)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"],
                                    textColor=CONDUIT_BLUE, fontSize=11, spaceBefore=12, spaceAfter=4)
    note_style = ParagraphStyle("Note", parent=styles["Normal"],
                                 fontSize=8, textColor=colors.orange)

    story = []

    # Header
    story.append(Paragraph(f"MEP Takeoff Report", title_style))
    story.append(Paragraph(f"{org_name}  ·  Job {job.id}  ·  {job.status.upper()}", subtitle_style))

    # Summary box
    total_cost = float(job.total_material_cost_usd or 0)
    summary_data = [
        ["Total Items", str(job.total_items or 0)],
        ["Total Material Cost", f"${total_cost:,.2f}"],
        ["AI Model", job.model_version or "N/A"],
        ["Actual API Cost", f"${float(job.actual_cost_usd or 0):.4f}"],
        ["Requires Review", str(job.low_confidence_count or 0)],
    ]
    summary_table = Table(summary_data, colWidths=[2.5 * inch, 2.5 * inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D6E4F0")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.2 * inch))

    # Items table grouped by system
    story.append(Paragraph("Component Detail", section_style))

    col_headers = ["Type", "Tag", "Qty", "Unit", "Specification", "System", "Unit $", "Total $"]
    col_widths_pdf = [0.8*inch, 0.7*inch, 0.5*inch, 0.5*inch, 2.5*inch, 0.8*inch, 0.7*inch, 0.8*inch]

    table_data = [col_headers]
    row_colors = []
    row_num = 1

    systems: dict[str, list[Any]] = {}
    for item in items:
        s = item.system or "Other"
        systems.setdefault(s, []).append(item)

    for system, sys_items in sorted(systems.items()):
        # System group header
        table_data.append([system.upper(), "", "", "", "", "", "", ""])
        row_colors.append(("BACKGROUND", (0, row_num), (-1, row_num), colors.HexColor("#1F4E79")))
        row_colors.append(("TEXTCOLOR", (0, row_num), (-1, row_num), colors.white))
        row_colors.append(("FONTNAME", (0, row_num), (-1, row_num), "Helvetica-Bold"))
        row_num += 1

        for item in sys_items:
            row = [
                item.type,
                item.tag or "",
                f"{float(item.quantity):.1f}",
                item.unit,
                item.specification[:45],
                item.system or "",
                f"${float(item.unit_cost_usd):.2f}" if item.unit_cost_usd else "",
                f"${float(item.total_cost_usd):.2f}" if item.total_cost_usd else "",
            ]
            table_data.append(row)
            if item.requires_review:
                row_colors.append(("BACKGROUND", (0, row_num), (-1, row_num), colors.HexColor("#FFF2CC")))
            row_num += 1

    items_table = Table(table_data, colWidths=col_widths_pdf)
    base_style = [
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("BACKGROUND", (0, 0), (-1, 0), CONDUIT_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    base_style.extend(row_colors)
    items_table.setStyle(TableStyle(base_style))
    story.append(items_table)

    # Low-confidence notes
    low_conf = [i for i in items if i.requires_review]
    if low_conf:
        story.append(Spacer(1, 0.15 * inch))
        story.append(Paragraph("⚠ Items Requiring Field Verification", section_style))
        for item in low_conf:
            story.append(Paragraph(
                f"• {item.type} — {item.specification[:60]} (confidence: {item.confidence}%)"
                + (f" — {item.notes}" if item.notes else ""),
                note_style,
            ))

    # Approver signature
    story.append(Spacer(1, 0.3 * inch))
    sig_data = [
        ["Approved by:", approver_name or "________________________"],
        ["Date:", "________________________"],
        ["Signature:", "________________________"],
    ]
    sig_table = Table(sig_data, colWidths=[1.5 * inch, 3 * inch])
    sig_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(sig_table)

    doc.build(story)
    return buf.getvalue()
